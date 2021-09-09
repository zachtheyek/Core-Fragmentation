import matplotlib.pyplot as plt
import openpyxl as opxl
from pathlib import Path

#Set the path to the excel file using Path()
spreadsheet_path = Path("Initial Detections.xlsx")

#Then, read in the excel file using opxl.load_workbook()
spreadsheet = opxl.load_workbook(spreadsheet_path, data_only = True) 

#Read the active sheet using the .actve method
sheet = spreadsheet.active


dirty_SN, selfcal_SN = [], []
coln = 0

for column in sheet.iter_cols(min_row = 2, max_col = 10, max_row = 101):
    coln += 1

    if coln == 6:
        for cell in column:
            if cell.value:
                dirty_SN.append(cell.value)

    elif coln == 9:
        for cell in column:
            if cell.value:
                selfcal_SN.append(cell.value)

imprv_factor = [i / j for i, j in zip(selfcal_SN, dirty_SN)] 

plt.scatter(dirty_SN, imprv_factor)
plt.xlim([0., 100.])
plt.ylim([0., 8.])
plt.xlabel("Dirty Signal-to-Noise Ratio")
plt.ylabel("Improvement Factor (Post-Selfcal)")
plt.savefig("improvement_factor.eps")
plt.show()
